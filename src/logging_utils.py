"""Processing evidence: a persistent, append-only CSV step log and per-output-folder
processing_log.xlsx summaries.

Logging is embedded directly in the processing functions via an injected logger
(defaulting to NULL_LOGGER, a no-op), so calling those functions in isolation - as
the unit tests do - produces no log output unless a real logger is passed in.

Never log password/secret values - only booleans/flags about whether one was needed.
"""

import csv
import uuid
from datetime import datetime
from pathlib import Path

import openpyxl

_CSV_FIELDNAMES = [
    "run_id", "timestamp", "source_file", "ingestion_type", "worksheet_name", "step", "status", "detail",
]


def _now_display():
    """Human-readable local timestamp (dd/mm/yyyy hh:mm)."""
    return datetime.now().strftime("%d/%m/%Y %H:%M")


class NullProcessingLogger:
    """No-op logger - the default for processing functions."""

    def log_step(self, source_file, ingestion_type, worksheet_name, step, status, detail=""):
        pass


NULL_LOGGER = NullProcessingLogger()


class ProcessingLogger:
    """Appends one row per processing step to a single persistent CSV log that
    accumulates across every run (the file is never replaced or rotated)."""

    def __init__(self, csv_path, run_id):
        self.csv_path = Path(csv_path)
        self.run_id = run_id

    @classmethod
    def start_run(cls, csv_path):
        csv_path = Path(csv_path)
        csv_path.parent.mkdir(parents=True, exist_ok=True)
        if not csv_path.exists():
            with open(csv_path, "w", newline="", encoding="utf-8") as f:
                csv.writer(f).writerow(_CSV_FIELDNAMES)

        logger = cls(csv_path, uuid.uuid4().hex)
        logger.log_step("", "", "", "run_start", "ok", "")
        return logger

    def log_step(self, source_file, ingestion_type, worksheet_name, step, status, detail=""):
        row = [
            self.run_id, _now_display(), str(source_file), ingestion_type or "",
            worksheet_name or "", step, status, detail,
        ]
        with open(self.csv_path, "a", newline="", encoding="utf-8") as f:
            csv.writer(f).writerow(row)

    def log_run_end(self):
        self.log_step("", "", "", "run_end", "ok", "")


# ---------------------------------------------------------------------------
# Per-output-folder processing_log.xlsx (rebuilt fresh each run, latest run only)
# ---------------------------------------------------------------------------

def build_processing_log_workbook(file_summaries, worksheet_results):
    """Build the 3-sheet processing_log.xlsx for one ingestion type's output folder,
    covering only the worksheets/files processed in the run that just completed.
    """
    workbook = openpyxl.Workbook()
    workbook.remove(workbook.active)

    _build_granular_sheet(workbook, worksheet_results)
    _build_overview_sheet(workbook, file_summaries, worksheet_results)
    _build_column_mapping_sheet(workbook, worksheet_results)

    return workbook


def _build_granular_sheet(workbook, worksheet_results):
    ws = workbook.create_sheet("Granular")
    ws.append([
        "Source File", "Worksheet", "Status", "Anchor Pair Used", "Rows Extracted",
        "Columns Extracted", "Validation Passed", "Failed Rules", "Output Location",
        "Contract Code Year", "Warnings", "Errors",
    ])
    for result in worksheet_results:
        header_match = result.header_match
        validation = result.validation
        failed_rules = (
            ", ".join(r.name for r in validation.rule_results if not r.passed) if validation else ""
        )
        ws.append([
            str(result.source_file),
            result.worksheet_name,
            result.status,
            "/".join(header_match.matched_pair) if header_match else "",
            result.data.row_count if result.data else 0,
            (header_match.end_col - header_match.start_col + 1) if header_match else 0,
            validation.passed if validation else "",
            failed_rules,
            str(result.output_path) if result.output_path else "",
            result.contract_code_year or "",
            "; ".join(result.warnings),
            "; ".join(result.errors),
        ])


def _build_overview_sheet(workbook, file_summaries, worksheet_results):
    ws = workbook.create_sheet("Overview")
    ws.append(["Metric", "Count"])

    status_counts = {}
    for result in worksheet_results:
        status_counts[result.status] = status_counts.get(result.status, 0) + 1

    files_opened_ok = sum(1 for f in file_summaries if f["opened_ok"])

    rows = [
        ("Total files processed", len(file_summaries)),
        ("Files opened successfully", files_opened_ok),
        ("Files that failed to open", len(file_summaries) - files_opened_ok),
        ("Files with password protection", sum(1 for f in file_summaries if f["password_required"])),
        ("Worksheets extracted", status_counts.get("extracted", 0)),
        ("Worksheets skipped (no header)", status_counts.get("skipped_no_header", 0)),
        ("Worksheets skipped (no data)", status_counts.get("skipped_no_data", 0)),
        ("Worksheets errored", status_counts.get("error", 0)),
        ("Tables passed validation", sum(1 for r in worksheet_results if r.validation and r.validation.passed)),
        ("Tables failed validation", sum(1 for r in worksheet_results if r.validation and not r.validation.passed)),
    ]
    for metric, count in rows:
        ws.append([metric, count])


def _build_column_mapping_sheet(workbook, worksheet_results):
    """Source = distinct column names seen in extracted headers; Frequency = number of
    distinct (file, sheet) pairs that column name appeared in; Sink is left blank for
    manual completion.
    """
    ws = workbook.create_sheet("ColumnMapping")
    ws.append(["Source", "Frequency", "Sink"])

    frequency = {}
    for result in worksheet_results:
        if result.status != "extracted" or result.header_match is None:
            continue
        distinct_columns_in_sheet = {v for v in result.header_match.header_values if v is not None}
        for column_name in distinct_columns_in_sheet:
            frequency[column_name] = frequency.get(column_name, 0) + 1

    for column_name, count in sorted(frequency.items(), key=lambda item: item[1], reverse=True):
        ws.append([column_name, count, None])
