"""Per-worksheet extraction orchestration and output writing.

Ties anchor_detection + metadata_extraction together per worksheet, and (per the
approved plan) also owns output writing - the mandated module list has no separate
output module, so this is its natural home.
"""

import re
from dataclasses import dataclass, field
from pathlib import Path

import openpyxl

from src.anchor_detection import find_header_row
from src.contract_detection import detect_contract_code_year
from src.logging_utils import NULL_LOGGER
from src.metadata_extraction import extract_metadata, join_metadata

_MONTH_NAMES = (
    r"jan(?:uary)?|feb(?:ruary)?|mar(?:ch)?|apr(?:il)?|may|jun(?:e)?|jul(?:y)?"
    r"|aug(?:ust)?|sep(?:t(?:ember)?)?|oct(?:ober)?|nov(?:ember)?|dec(?:ember)?"
)
_DATE_PATTERN = re.compile(
    r"\d{4}-\d{2}(?:-\d{2})?"                          # numeric ISO: 2026-07, 2026-07-15
    r"|\b(?:" + _MONTH_NAMES + r")\b[\s_-]*\d{4}"       # month name + year: Aug 2026, August-2026
    r"|\d{4}[\s_-]*\b(?:" + _MONTH_NAMES + r")\b",      # year + month name: 2026 Aug
    re.IGNORECASE,
)
_MAX_SHEET_NAME_LEN = 31


@dataclass
class DataBlock:
    rows: list
    row_count: int
    row_numbers: list = field(default_factory=list)  # source worksheet row per entry in rows


@dataclass
class WorksheetResult:
    source_file: Path
    ingestion_type: str
    worksheet_name: str
    status: str  # "extracted" | "skipped_no_header" | "skipped_no_data" | "error"
    header_match: object = None
    metadata: object = None
    data: object = None
    output_path: Path = None
    validation: object = None
    contract_code_year: str = None
    contract_section: str = None
    warnings: list = field(default_factory=list)
    errors: list = field(default_factory=list)


# ---------------------------------------------------------------------------
# Extraction
# ---------------------------------------------------------------------------

def process_worksheet(
    worksheet, worksheet_name, source_file, ingestion_type, anchor_pairs, contract_config, logger=NULL_LOGGER,
):
    """Detect the header, extract metadata + data, classify the worksheet's status, and
    (for extracted tables) detect the ContractCodeYear. Logs the anchor_detection,
    metadata_extraction, data_extraction and contract_detection steps.
    """
    header_match = find_header_row(worksheet, anchor_pairs)
    if header_match is None:
        logger.log_step(source_file, ingestion_type, worksheet_name, "anchor_detection", "fail",
                         "no anchor pair matched any row")
        return WorksheetResult(
            source_file=source_file, ingestion_type=ingestion_type,
            worksheet_name=worksheet_name, status="skipped_no_header",
        )

    logger.log_step(
        source_file, ingestion_type, worksheet_name, "anchor_detection", "ok",
        f"matched {header_match.matched_pair[0]}/{header_match.matched_pair[1]} at row "
        f"{header_match.row_index} ({len(header_match.occurrence_rows)} occurrence(s))",
    )

    metadata = extract_metadata(worksheet, header_match.row_index)
    logger.log_step(source_file, ingestion_type, worksheet_name, "metadata_extraction", "ok",
                     f"{len(metadata.values)} metadata values extracted")

    data = extract_data_block(worksheet, header_match)
    meaningful = has_meaningful_data(data)
    logger.log_step(source_file, ingestion_type, worksheet_name, "data_extraction",
                     "ok" if meaningful else "fail", f"{data.row_count} data rows extracted")

    if not meaningful:
        return WorksheetResult(
            source_file=source_file, ingestion_type=ingestion_type,
            worksheet_name=worksheet_name, status="skipped_no_data",
            header_match=header_match, metadata=metadata, data=data,
        )

    detection = detect_contract_code_year(
        source_file=source_file, worksheet_name=worksheet_name, metadata=metadata,
        header_values=header_match.header_values, data_rows=data.rows,
        header_row=header_match.row_index, start_col=header_match.start_col, data_row_numbers=data.row_numbers,
        contract_config=contract_config, logger=logger, ingestion_type=ingestion_type,
    )
    logger.log_step(
        source_file, ingestion_type, worksheet_name, "contract_detection",
        "ok" if detection.status == "resolved" else detection.status, detection.detail,
    )

    result = WorksheetResult(
        source_file=source_file, ingestion_type=ingestion_type,
        worksheet_name=worksheet_name, status="extracted",
        header_match=header_match, metadata=metadata, data=data,
        contract_code_year=detection.contract_code_year, contract_section=detection.section,
    )
    if detection.status != "resolved":
        result.warnings.append(detection.detail)
    return result


def extract_data_block(worksheet, header_match):
    """Extract data rows below the header, applying the row-inclusion filter and
    fully-blank termination rule (see CLAUDE.md-approved plan §6) independently for
    every occurrence of the header, concatenating the results in occurrence order.
    """
    anchor_cols = sorted(set(header_match.matched_pair_cols.values()))
    start_col, end_col = header_match.start_col, header_match.end_col

    rows = []
    row_numbers = []
    for occurrence_row in header_match.occurrence_rows:
        block_rows, block_row_numbers = _extract_block_rows(worksheet, occurrence_row, start_col, end_col, anchor_cols)
        rows.extend(block_rows)
        row_numbers.extend(block_row_numbers)

    return DataBlock(rows=rows, row_count=len(rows), row_numbers=row_numbers)


def _extract_block_rows(worksheet, header_row, start_col, end_col, anchor_cols):
    rows = []
    row_numbers = []
    row_index = header_row + 1
    max_row = worksheet.max_row
    while row_index <= max_row:
        row_values = [worksheet.cell(row=row_index, column=c).value for c in range(start_col, end_col + 1)]
        if all(v is None for v in row_values):
            break  # fully-blank row: end of this block

        anchor_values = [worksheet.cell(row=row_index, column=c).value for c in anchor_cols]
        if all(v is not None for v in anchor_values):
            rows.append(row_values)
            row_numbers.append(row_index)
        # else: row skipped (blank anchor columns), scan continues

        row_index += 1
    return rows, row_numbers


def has_meaningful_data(data_block):
    return data_block.row_count > 0


# ---------------------------------------------------------------------------
# Output writing
# ---------------------------------------------------------------------------

def worksheet_name_has_date(name):
    return bool(_DATE_PATTERN.search(name))


_RISK_FILENAME_KEYWORDS = ("rolling", "by month", "database")


def is_risk_filename(name):
    """Risk source files are split into risk/ vs premium/ output by filename: a name
    containing 'rolling', 'by month', or 'database' (case-insensitive, exact spelling)
    is Risk; anything else is Premium.
    """
    lowered = name.lower()
    return any(keyword in lowered for keyword in _RISK_FILENAME_KEYWORDS)


def dedupe_header(header_values):
    """First occurrence of a label is unchanged; each later duplicate gets a _N suffix."""
    seen_counts = {}
    result = []
    for value in header_values:
        if value is None:
            result.append(None)
            continue
        if value not in seen_counts:
            seen_counts[value] = 0
            result.append(value)
        else:
            seen_counts[value] += 1
            result.append(f"{value}_{seen_counts[value]}")
    return result


def _write_table(ws, header_values, data_rows, metadata_string, contract_code_year):
    ws.append(dedupe_header(header_values) + ["ContractCodeYear", "metadata"])
    for i, row in enumerate(data_rows):
        trailing = [contract_code_year, metadata_string] if i == 0 else [None, None]
        ws.append(list(row) + trailing)


def _unique_sheet_name(name, used_names):
    truncated = name[:_MAX_SHEET_NAME_LEN]
    candidate = truncated
    suffix = 1
    while candidate in used_names:
        suffix_str = f"_{suffix}"
        candidate = truncated[: _MAX_SHEET_NAME_LEN - len(suffix_str)] + suffix_str
        suffix += 1
    return candidate


def write_combined_workbook(output_path, worksheet_results, logger=NULL_LOGGER):
    """Write one sheet per result into a single workbook at output_path (overwriting
    any existing file). N=1 is the 'individual workbook' case - no separate code path
    is needed for that. Sets output_path on each result and logs an output_write step
    per result.
    """
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    workbook = openpyxl.Workbook()
    workbook.remove(workbook.active)

    used_sheet_names = set()
    for result in worksheet_results:
        sheet_name = _unique_sheet_name(result.worksheet_name, used_sheet_names)
        used_sheet_names.add(sheet_name)
        ws = workbook.create_sheet(title=sheet_name)

        metadata_string = join_metadata(result.metadata)
        _write_table(
            ws, result.header_match.header_values, result.data.rows, metadata_string,
            result.contract_code_year or "",
        )

        result.output_path = output_path
        logger.log_step(result.source_file, result.ingestion_type, result.worksheet_name,
                         "output_write", "ok", f"written to {output_path}")

    workbook.save(output_path)
