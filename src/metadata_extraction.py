"""Extraction of metadata content above a detected header row."""

from dataclasses import dataclass, field

from openpyxl.utils import get_column_letter

_CODE_ROW_PREFIX = "cr0"
_CODE_ROW_MIN_CELLS = 3


@dataclass
class MetadataBlock:
    values: list
    cells: list = field(default_factory=list)  # "A1"-style ref per value, same order/length as values


def _is_code_row(row_values):
    """A row of internal field codes (e.g. CR0013, CR0015, ...) sitting above the real
    metadata/header - not actual metadata, dropped like a blank row. Requires at least
    3 non-empty cells starting with 'CR0' (case-insensitive) rather than just one, so a
    row that only coincidentally has a single CR0-prefixed value isn't mistaken for a
    full code row.
    """
    matches = sum(
        1 for v in row_values if v is not None and str(v).strip().lower().startswith(_CODE_ROW_PREFIX)
    )
    return matches >= _CODE_ROW_MIN_CELLS


def extract_metadata(worksheet, header_row_index):
    """Flatten all non-empty cell values from rows 1..header_row_index-1 into reading
    order (row-major, top-to-bottom, then left-to-right). Blank cells, fully-blank
    rows, and internal code rows (see _is_code_row) are dropped. Not a fixed key/value
    pairing - rows may carry any number of non-empty cells.
    """
    values = []
    cells = []
    max_col = worksheet.max_column
    for row_index in range(1, header_row_index):
        row_values = [worksheet.cell(row=row_index, column=col).value for col in range(1, max_col + 1)]
        if _is_code_row(row_values):
            continue
        for col, value in enumerate(row_values, start=1):
            if value is not None:
                values.append(value)
                cells.append(f"{get_column_letter(col)}{row_index}")
    return MetadataBlock(values=values, cells=cells)


def join_metadata(metadata_block):
    """Join a MetadataBlock's flattened values into a single '|'-delimited string."""
    return "|".join(str(v) for v in metadata_block.values)
