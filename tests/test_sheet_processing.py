from pathlib import Path

import openpyxl
import pytest

from src.anchor_detection import HeaderMatch, find_header_row
from src.metadata_extraction import MetadataBlock
from src.sheet_processing import (
    DataBlock,
    WorksheetResult,
    dedupe_header,
    extract_data_block,
    has_meaningful_data,
    process_worksheet,
    worksheet_name_has_date,
    write_combined_workbook,
)


# ---------------------------------------------------------------------------
# extract_data_block: row-inclusion filter + fully-blank termination
# ---------------------------------------------------------------------------

def test_risk_main_skips_blank_anchor_row_but_keeps_row_after_it(risk_test_workbook, risk_anchor_pairs):
    ws = risk_test_workbook["Risk_Main"]
    match = find_header_row(ws, risk_anchor_pairs)
    data = extract_data_block(ws, match)

    assert data.row_count == 4
    policy_numbers = [row[0] for row in data.rows]
    assert policy_numbers == ["POL-1001", "POL-1002", "POL-1003", "POL-9999"]


def test_risk_altanchor_two_rows_then_fully_blank_row_terminates(risk_test_workbook, risk_anchor_pairs):
    ws = risk_test_workbook["Risk_AltAnchor"]
    match = find_header_row(ws, risk_anchor_pairs)
    data = extract_data_block(ws, match)

    assert data.row_count == 2
    assert [row[0] for row in data.rows] == ["POL-2001", "POL-2002"]


def test_risk_nodata_yields_zero_rows(risk_test_workbook, risk_anchor_pairs):
    ws = risk_test_workbook["Risk_NoData"]
    match = find_header_row(ws, risk_anchor_pairs)
    data = extract_data_block(ws, match)

    assert data.row_count == 0
    assert has_meaningful_data(data) is False


def test_risk_main1_concatenates_both_occurrences(risk_main1_workbook, risk_anchor_pairs):
    ws = risk_main1_workbook["Risk_Main1"]
    match = find_header_row(ws, risk_anchor_pairs)
    data = extract_data_block(ws, match)

    assert data.row_count == 8
    policy_numbers = [row[0] for row in data.rows]
    assert policy_numbers == [
        "POL-1001", "POL-1002", "POL-1003", "POL-9999",
        "POL-1004", "POL-1005", "POL-1006", "POL-11111",
    ]


def test_claims_main_same_skip_include_pattern(claims_test_workbook, claims_anchor_pairs):
    ws = claims_test_workbook["Claims_Main"]
    match = find_header_row(ws, claims_anchor_pairs)
    data = extract_data_block(ws, match)

    assert data.row_count == 4
    assert [row[0] for row in data.rows] == ["CLM-001", "CLM-002", "CLM-003", "CLM-999"]


def test_claims_secondary_three_rows(claims_test_workbook, claims_anchor_pairs):
    ws = claims_test_workbook["Claims_Secondary"]
    match = find_header_row(ws, claims_anchor_pairs)
    data = extract_data_block(ws, match)

    assert data.row_count == 3
    assert [row[0] for row in data.rows] == ["POL-3001", "POL-3002", "POL-3003"]


# ---------------------------------------------------------------------------
# process_worksheet: status classification
# ---------------------------------------------------------------------------

def test_process_worksheet_extracted(risk_test_workbook, risk_anchor_pairs):
    ws = risk_test_workbook["Risk_Main"]
    result = process_worksheet(ws, "Risk_Main", Path("file1_Risk_Input.xlsx"), "risk", risk_anchor_pairs)

    assert result.status == "extracted"
    assert result.data.row_count == 4


def test_process_worksheet_skipped_no_data(risk_test_workbook, risk_anchor_pairs):
    ws = risk_test_workbook["Risk_NoData"]
    result = process_worksheet(ws, "Risk_NoData", Path("file1_Risk_Input.xlsx"), "risk", risk_anchor_pairs)

    assert result.status == "skipped_no_data"


def test_process_worksheet_skipped_no_header(risk_test_workbook, risk_anchor_pairs):
    ws = risk_test_workbook["Notes"]
    result = process_worksheet(ws, "Notes", Path("file1_Risk_Input.xlsx"), "risk", risk_anchor_pairs)

    assert result.status == "skipped_no_header"


# ---------------------------------------------------------------------------
# dedupe_header
# ---------------------------------------------------------------------------

def test_dedupe_header_suffixes_duplicates():
    assert dedupe_header(["Policy Number", "Premium", "Premium", "Region"]) == [
        "Policy Number", "Premium", "Premium_1", "Region",
    ]


def test_dedupe_header_handles_three_duplicates():
    assert dedupe_header(["X", "X", "X"]) == ["X", "X_1", "X_2"]


def test_dedupe_header_leaves_unique_values_and_none_unchanged():
    assert dedupe_header(["A", None, "B"]) == ["A", None, "B"]


# ---------------------------------------------------------------------------
# worksheet_name_has_date
# ---------------------------------------------------------------------------

@pytest.mark.parametrize("name,expected", [
    ("Claims_Main", False),
    ("Claims_Secondary", False),
    ("ReadMe", False),
    ("Claims_2026-07", True),
    ("Report_2026-08-01", True),
    ("Claims_Main Aug 2026", True),
    ("August-2026 Claims", True),
    ("2026 Aug Claims", True),
    ("Claims_Main1", False),  # not a false positive on "may"/"jun"-style substrings
])
def test_worksheet_name_has_date(name, expected):
    assert worksheet_name_has_date(name) is expected


# ---------------------------------------------------------------------------
# write_combined_workbook: grouping, metadata column, header dedup, overwrite
# ---------------------------------------------------------------------------

def _make_result(worksheet_name, header_values, matched_pair_cols, data_rows, metadata_values):
    header_match = HeaderMatch(
        row_index=1, start_col=1, end_col=len(header_values), header_values=header_values,
        matched_pair=("A", "B"), matched_pair_cols=matched_pair_cols, occurrence_rows=[1],
    )
    data = DataBlock(rows=data_rows, row_count=len(data_rows))
    metadata = MetadataBlock(values=metadata_values)
    return WorksheetResult(
        source_file=Path("src.xlsx"), ingestion_type="risk", worksheet_name=worksheet_name,
        status="extracted", header_match=header_match, metadata=metadata, data=data,
    )


def test_write_combined_workbook_single_result_is_one_sheet(tmp_path):
    result = _make_result(
        "Risk_Main", ["Policy Number", "Premium", "Premium"], {"A": 1, "B": 2},
        [["POL-1", 100, 10], ["POL-2", 200, 20]], ["Broker", "Acme"],
    )
    output_path = tmp_path / "out.xlsx"

    write_combined_workbook(output_path, [result])

    assert result.output_path == output_path
    workbook = openpyxl.load_workbook(output_path)
    assert workbook.sheetnames == ["Risk_Main"]
    ws = workbook["Risk_Main"]
    assert [c.value for c in ws[1]] == ["Policy Number", "Premium", "Premium_1", "metadata"]
    assert [c.value for c in ws[2]] == ["POL-1", 100, 10, "Broker|Acme"]
    assert [c.value for c in ws[3]] == ["POL-2", 200, 20, None]


def test_write_combined_workbook_multiple_results_multiple_sheets(tmp_path):
    result_a = _make_result("Risk_Main", ["A"], {"x": 1}, [["v1"]], ["m1"])
    result_b = _make_result("Risk_AltAnchor", ["A"], {"x": 1}, [["v2"]], ["m2"])
    output_path = tmp_path / "combined.xlsx"

    write_combined_workbook(output_path, [result_a, result_b])

    workbook = openpyxl.load_workbook(output_path)
    assert workbook.sheetnames == ["Risk_Main", "Risk_AltAnchor"]
    assert result_a.output_path == output_path
    assert result_b.output_path == output_path


def test_write_combined_workbook_overwrites_existing_file(tmp_path):
    output_path = tmp_path / "out.xlsx"
    output_path.write_bytes(b"not a real workbook")

    result = _make_result("Risk_Main", ["A"], {"x": 1}, [["v1"]], [])
    write_combined_workbook(output_path, [result])

    workbook = openpyxl.load_workbook(output_path)
    assert workbook.sheetnames == ["Risk_Main"]
