from pathlib import Path

import openpyxl
import pytest

from src.contract_detection import load_contract_codes

TEST_FILES_DIR = Path(__file__).resolve().parent / "test files"
CONTRACT_CODES_PATH = Path(__file__).resolve().parent.parent / "config" / "contract_codes.json"


def _load_sheet(filename, sheet_name):
    workbook = openpyxl.load_workbook(TEST_FILES_DIR / filename, data_only=True)
    return workbook[sheet_name]


@pytest.fixture
def risk_test_workbook():
    return openpyxl.load_workbook(TEST_FILES_DIR / "Hardy_Risk_test.xlsx", data_only=True)


@pytest.fixture
def risk_main1_workbook():
    return openpyxl.load_workbook(TEST_FILES_DIR / "file1_Risk_test1.xlsx", data_only=True)


@pytest.fixture
def claims_test_workbook():
    return openpyxl.load_workbook(TEST_FILES_DIR / "HDI_Claims_test.xlsx", data_only=True)


@pytest.fixture
def sections_claims_workbook():
    return openpyxl.load_workbook(TEST_FILES_DIR / "file_Claims_test.xlsx", data_only=True)


@pytest.fixture
def contract_config():
    return load_contract_codes(CONTRACT_CODES_PATH)


@pytest.fixture
def risk_anchor_pairs():
    return [
        ("Policy Number", "Inception Date"),
        ("Policy Number External", "Sum Insured"),
        ("Broker", "Effective Date"),
    ]


@pytest.fixture
def claims_anchor_pairs():
    return [
        ("Claim Number", "Claim Amount"),
        ("Insured Name", "Paid Amount"),
        ("Policy Number", "Date Reported"),
        ("Policy Number Source", "Effective Date Value"),
    ]
