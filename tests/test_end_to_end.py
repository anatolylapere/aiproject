import csv
from pathlib import Path

import openpyxl

import main
from src.anchor_detection import load_anchor_pairs
from src.contract_detection import load_contract_codes
from src.file_access import FileListEntry
from src.logging_utils import ProcessingLogger

TEST_FILES_DIR = Path(__file__).resolve().parent / "test files"
CONFIG_PATH = Path(__file__).resolve().parent.parent / "config" / "anchor_pairs.json"
CONTRACT_CODES_PATH = Path(__file__).resolve().parent.parent / "config" / "contract_codes.json"


def _run(monkeypatch, tmp_path, entries):
    output_root = tmp_path / "output"
    logs_dir = tmp_path / "logs"
    monkeypatch.setattr(main, "OUTPUT_ROOT", output_root)
    monkeypatch.setattr(main, "LOGS_DIR", logs_dir)

    anchors = load_anchor_pairs(CONFIG_PATH)
    contract_config = load_contract_codes(CONTRACT_CODES_PATH)
    logger = ProcessingLogger.start_run(logs_dir / "processing_log.csv")

    worksheet_results_by_type = {"risk": [], "claims": []}
    file_summaries_by_type = {"risk": [], "claims": []}
    for entry in entries:
        worksheet_results, file_summary = main.process_entry(entry, anchors, contract_config, logger)
        worksheet_results_by_type[entry.ingestion_type].extend(worksheet_results)
        file_summaries_by_type[entry.ingestion_type].append(file_summary)

    for ingestion_type in ("risk", "claims"):
        main.write_processing_log_workbook(
            ingestion_type, file_summaries_by_type[ingestion_type], worksheet_results_by_type[ingestion_type],
        )
    logger.log_run_end()

    return output_root, logger.csv_path, worksheet_results_by_type


def _read_csv_rows(csv_path):
    with open(csv_path, newline="", encoding="utf-8") as f:
        return list(csv.DictReader(f))


def _statuses(worksheet_results):
    return {r.worksheet_name: r.status for r in worksheet_results}


def test_risk_fixture_combines_into_one_workbook(tmp_path, monkeypatch):
    entry = FileListEntry(source_file=TEST_FILES_DIR / "Hardy_Risk_test.xlsx", ingestion_type="risk")

    output_root, csv_path, results_by_type = _run(monkeypatch, tmp_path, [entry])

    # both sheets resolve the same parent code (BW01972), so the combined Risk file is
    # nested under a BW01972/ folder with the code appended to its filename too. The
    # filename has none of the risk keywords ("rolling"/"by month"/"database"), so it
    # routes to output/premium/, not output/risk/.
    output_file = output_root / "premium" / "BW01972" / "Hardy_Risk_test__BW01972.xlsx"
    assert output_file.exists()
    workbook = openpyxl.load_workbook(output_file)
    assert set(workbook.sheetnames) == {"Risk_Main", "Risk_AltAnchor"}

    assert _statuses(results_by_type["risk"]) == {
        "Risk_Main": "extracted",
        "Risk_AltAnchor": "extracted",
        "Risk_NoData": "skipped_no_data",
        "Notes": "skipped_no_header",
    }

    # ContractCodeYear resolved via the filename ("Hardy" -> BW01972); this contract
    # has sections defined, but neither sheet's name/metadata mentions Sec A/Sec B.
    risk_main = next(r for r in results_by_type["risk"] if r.worksheet_name == "Risk_Main")
    assert risk_main.contract_code_year == "BW01972"

    # step-level CSV log carries granular evidence, not just a final status
    rows = _read_csv_rows(csv_path)
    anchor_steps = [r for r in rows if r["step"] == "anchor_detection" and r["worksheet_name"] == "Risk_Main"]
    assert anchor_steps[0]["status"] == "ok"
    assert "Policy Number/Inception Date" in anchor_steps[0]["detail"]

    contract_steps = [r for r in rows if r["step"] == "contract_detection" and r["worksheet_name"] == "Risk_Main"]
    assert contract_steps[0]["status"] == "ok"
    assert "BW01972" in contract_steps[0]["detail"]


def test_risk_filename_with_rolling_keyword_routes_to_risk_folder(tmp_path, monkeypatch):
    # Synthetic fixture: a Risk source file whose name contains "Rolling" - proves the
    # risk/ path independently of the real fixtures below (also named with "Rolling").
    source_path = tmp_path / "input" / "file_Risk_Hardy Rolling Total.xlsx"
    source_path.parent.mkdir(parents=True)
    workbook = openpyxl.Workbook()
    workbook.remove(workbook.active)

    ws = workbook.create_sheet("Risk_Main")
    ws.append(["Policy Number", "Inception Date"])
    ws.append(["POL-1", "2026-01-01"])

    workbook.save(source_path)

    entry = FileListEntry(source_file=source_path, ingestion_type="risk")
    output_root, _, results_by_type = _run(monkeypatch, tmp_path, [entry])

    risk_main = next(r for r in results_by_type["risk"] if r.worksheet_name == "Risk_Main")
    assert risk_main.contract_code_year == "BW01972"  # "Hardy" in the filename

    output_file = output_root / "risk" / "BW01972" / "file_Risk_Hardy Rolling Total__BW01972.xlsx"
    assert output_file.exists()
    assert not (output_root / "premium").exists()


def test_risk_property_sec_column_splits_sheets_by_section(tmp_path, monkeypatch):
    # file1_Risk_ HDI Rolling Total.xlsx: filename resolves BW01973 (HDI); Risk_Main's
    # Property Sec column reads "A" for every row, Risk_AltAnchor's reads "B" - each
    # sheet resolves a different section, so unlike a plain combined Risk file, they
    # must each land in their own output file (Risk now splits by section like Claims).
    entry = FileListEntry(source_file=TEST_FILES_DIR / "file1_Risk_ HDI Rolling Total.xlsx", ingestion_type="risk")

    output_root, _, results_by_type = _run(monkeypatch, tmp_path, [entry])

    statuses = {r.worksheet_name: r.contract_code_year for r in results_by_type["risk"]}
    assert statuses["Risk_Main"] == "BW01973A"
    assert statuses["Risk_AltAnchor"] == "BW01973B"

    # filename correctly spells "Rolling", so this routes to risk/, not premium/
    # (processing_log.xlsx is always written per ingestion type, alongside the data)
    risk_dir = output_root / "risk"
    assert sorted(p.name for p in risk_dir.iterdir()) == sorted(["BW01973A", "BW01973B", "processing_log.xlsx"])
    assert not (output_root / "premium").exists()

    section_a_file = openpyxl.load_workbook(
        risk_dir / "BW01973A" / "file1_Risk_ HDI Rolling Total__BW01973A.xlsx"
    )
    assert section_a_file.sheetnames == ["Risk_Main"]

    section_b_file = openpyxl.load_workbook(
        risk_dir / "BW01973B" / "file1_Risk_ HDI Rolling Total__BW01973B.xlsx"
    )
    assert section_b_file.sheetnames == ["Risk_AltAnchor"]


def test_claims_fixture_combines_since_no_sheet_name_has_a_date(tmp_path, monkeypatch):
    entry = FileListEntry(source_file=TEST_FILES_DIR / "HDI_Claims_test.xlsx", ingestion_type="claims")

    output_root, _, results_by_type = _run(monkeypatch, tmp_path, [entry])

    # both sheets resolve the same parent code (BW01973), so the combined file is
    # nested under a BW01973/ folder with the code appended to its filename too
    combined_file = output_root / "claims" / "BW01973" / "HDI_Claims_test__BW01973.xlsx"
    assert combined_file.exists()
    workbook = openpyxl.load_workbook(combined_file)
    assert set(workbook.sheetnames) == {"Claims_Main", "Claims_Secondary"}

    claims_dir = output_root / "claims"
    assert sorted(p.name for p in claims_dir.iterdir()) == sorted(["BW01973", "processing_log.xlsx"])
    assert sorted(p.name for p in (claims_dir / "BW01973").iterdir()) == ["HDI_Claims_test__BW01973.xlsx"]

    assert _statuses(results_by_type["claims"]) == {
        "Claims_Main": "extracted",
        "Claims_Secondary": "extracted",
        "Claims_NoData": "skipped_no_data",
        "ReadMe": "skipped_no_header",
    }

    # ContractCodeYear resolved via filename ("HDI" -> BW01973); no sections defined
    claims_main = next(r for r in results_by_type["claims"] if r.worksheet_name == "Claims_Main")
    assert claims_main.contract_code_year == "BW01973"


def test_claims_month_name_date_in_sheet_name_gets_its_own_file(tmp_path, monkeypatch):
    # Synthetic fixture: one sheet named with a month-name-style date ("Aug 2026", not
    # the numeric ISO format), one without - exercises worksheet_name_has_date's
    # month-name regex path end-to-end. Built in-memory (like
    # test_claims_worksheet_with_date_in_name_gets_its_own_file's ISO-date case above)
    # rather than depending on a shared fixture file, since file3_Claims_test.xlsx is
    # now dedicated to contract-detection section-fixture testing.
    source_path = tmp_path / "input" / "Claims_MonthName.xlsx"
    source_path.parent.mkdir(parents=True)
    workbook = openpyxl.Workbook()
    workbook.remove(workbook.active)

    dated_sheet = workbook.create_sheet("Claims_Main Aug 2026")
    dated_sheet.append(["Claim Number", "Claim Amount"])
    dated_sheet.append(["CLM-001", 2500])
    dated_sheet.append(["CLM-002", 7850])

    no_date_sheet = workbook.create_sheet("Claims_Secondary")
    no_date_sheet.append(["Claim Number", "Claim Amount"])
    no_date_sheet.append(["CLM-101", 12000])

    workbook.save(source_path)

    entry = FileListEntry(source_file=source_path, ingestion_type="claims")
    output_root, _, results_by_type = _run(monkeypatch, tmp_path, [entry])

    claims_dir = output_root / "claims"
    own_file = claims_dir / "Claims_MonthName__Claims_Main Aug 2026.xlsx"
    combined_file = claims_dir / "Claims_MonthName.xlsx"

    assert own_file.exists()
    assert combined_file.exists()

    own_workbook = openpyxl.load_workbook(own_file)
    assert own_workbook.sheetnames == ["Claims_Main Aug 2026"]

    combined_workbook = openpyxl.load_workbook(combined_file)
    assert combined_workbook.sheetnames == ["Claims_Secondary"]

    assert _statuses(results_by_type["claims"]) == {
        "Claims_Main Aug 2026": "extracted",
        "Claims_Secondary": "extracted",
    }


def test_claims_sections_fixture_resolves_base_contract_and_section(tmp_path, monkeypatch):
    # file_Claims_test.xlsx: sheets named "Section A"/"Section B". The base contract is
    # embedded directly in the metadata's "broker ref" value ("b1262bw0197219" contains
    # "bw01972" with no separators) - a code-style alias match doesn't require boundary
    # isolation, so it resolves BW01972, then the sheet name supplies the section letter.
    entry = FileListEntry(source_file=TEST_FILES_DIR / "file_Claims_test.xlsx", ingestion_type="claims")

    output_root, csv_path, results_by_type = _run(monkeypatch, tmp_path, [entry])

    section_a = next(r for r in results_by_type["claims"] if r.worksheet_name == "Section A")
    assert section_a.status == "extracted"
    assert section_a.contract_code_year == "BW01972A"
    assert section_a.warnings == []

    # section-bearing worksheets each get their own file, named after ContractCodeYear
    # and nested under a folder of that same name - not merged together, and not
    # merged with the unresolved Claims_Secondary sheet (which has no code, so it stays
    # flat at the claims/ root).
    claims_dir = output_root / "claims"
    assert sorted(p.name for p in claims_dir.iterdir()) == sorted([
        "BW01972A", "BW01972B", "file_Claims_test.xlsx", "processing_log.xlsx",
    ])

    section_a_file = openpyxl.load_workbook(claims_dir / "BW01972A" / "file_Claims_test__BW01972A.xlsx")
    assert section_a_file.sheetnames == ["Section A"]
    ws = section_a_file["Section A"]
    header = [c.value for c in ws[1]]
    assert header[-2:] == ["ContractCodeYear", "metadata"]
    assert [c.value for c in ws[2]][-2] == "BW01972A"

    section_b_file = openpyxl.load_workbook(claims_dir / "BW01972B" / "file_Claims_test__BW01972B.xlsx")
    assert section_b_file.sheetnames == ["Section B"]

    combined_file = openpyxl.load_workbook(claims_dir / "file_Claims_test.xlsx")
    assert combined_file.sheetnames == ["Claims_Secondary"]

    rows = _read_csv_rows(csv_path)
    contract_steps = [r for r in rows if r["step"] == "contract_detection" and r["worksheet_name"] == "Section A"]
    assert contract_steps[0]["status"] == "ok"
    assert "BW01972" in contract_steps[0]["detail"] and "section A" in contract_steps[0]["detail"]


def test_claims_second_sections_fixture_also_splits_by_contract_code_year(tmp_path, monkeypatch):
    # file3_Claims_test.xlsx: sheets named "Section B"/"SecA" (different naming style
    # than file_Claims_test.xlsx's "Section A"/"Section B", same underlying rule).
    # Claims_Secondary also carries the embedded code in its metadata but no section,
    # so its combined-file name gets the parent code appended too - every sheet in this
    # fixture now resolves a contract, so every output file has a code in its name.
    entry = FileListEntry(source_file=TEST_FILES_DIR / "file3_Claims_test.xlsx", ingestion_type="claims")

    output_root, _, results_by_type = _run(monkeypatch, tmp_path, [entry])

    claims_secondary = next(r for r in results_by_type["claims"] if r.worksheet_name == "Claims_Secondary")
    assert claims_secondary.contract_code_year == "BW01972"
    assert claims_secondary.contract_section is None

    # every sheet resolves a code, so every output lands under its own BW01972*/ folder
    claims_dir = output_root / "claims"
    assert sorted(p.name for p in claims_dir.iterdir()) == sorted([
        "BW01972A", "BW01972B", "BW01972", "processing_log.xlsx",
    ])

    sec_a_file = openpyxl.load_workbook(claims_dir / "BW01972A" / "file3_Claims_test__BW01972A.xlsx")
    assert sec_a_file.sheetnames == ["SecA"]

    sec_b_file = openpyxl.load_workbook(claims_dir / "BW01972B" / "file3_Claims_test__BW01972B.xlsx")
    assert sec_b_file.sheetnames == ["Section B"]

    combined_file = openpyxl.load_workbook(claims_dir / "BW01972" / "file3_Claims_test__BW01972.xlsx")
    assert combined_file.sheetnames == ["Claims_Secondary"]


def test_claims_code_row_above_header_is_ignored_as_metadata(tmp_path, monkeypatch):
    # Bw0197222 claims.xlsx: row 1 above the header is an internal "CR001..CR007" field
    # code row (dropped like a blank row - nothing else above the header, so metadata is
    # empty). Base resolves via the Agreement Number column (BW0197222 -> BW01972). The
    # sheet also has a "Loss Description" column with free text ("Legal expense Claim")
    # that doesn't coincidentally match any configured alias, so it doesn't affect the
    # result either way - see tests/test_contract_detection.py's
    # test_loss_description_column_excluded_from_evidence for a case where free text
    # WOULD cause a false ambiguity without the exclusion.
    entry = FileListEntry(source_file=TEST_FILES_DIR / "Bw0197222 claims.xlsx", ingestion_type="claims")

    output_root, _, results_by_type = _run(monkeypatch, tmp_path, [entry])

    claims_main = next(r for r in results_by_type["claims"] if r.worksheet_name == "Claims_Main")
    assert claims_main.contract_code_year == "BW01972"
    assert claims_main.metadata.values == []  # CR0 code row dropped, nothing else above the header
    assert claims_main.warnings == []  # resolved cleanly, no warning

    output_file = output_root / "claims" / "BW01972" / "Bw0197222 claims__BW01972.xlsx"
    assert output_file.exists()
    workbook = openpyxl.load_workbook(output_file)
    ws = workbook["Claims_Main"]
    header = [c.value for c in ws[1]]
    assert "Loss Description" in header  # still extracted/written - only excluded from evidence
    assert header[-1] == "metadata"
    assert [c.value for c in ws[2]][-1] is None  # no metadata string


def test_claims_worksheets_sharing_a_contract_code_year_share_one_file(tmp_path, monkeypatch):
    # Synthetic fixture: two sheets that both resolve to the same section (BW01972A) -
    # must land together in one output file, not overwrite each other.
    source_path = tmp_path / "input" / "Claims_SharedSection.xlsx"
    source_path.parent.mkdir(parents=True)
    workbook = openpyxl.Workbook()
    workbook.remove(workbook.active)

    for sheet_name in ("SecA_Batch1", "SecA_Batch2"):
        ws = workbook.create_sheet(sheet_name)
        ws.append(["Hardy", "SecA", None, None])
        ws.append(["Claim Number", "Claim Amount", None, None])
        ws.append([f"CLM-{sheet_name}", 100, None, None])

    workbook.save(source_path)

    entry = FileListEntry(source_file=source_path, ingestion_type="claims")
    output_root, _, results_by_type = _run(monkeypatch, tmp_path, [entry])

    statuses = {r.worksheet_name: r.contract_code_year for r in results_by_type["claims"]}
    assert statuses == {"SecA_Batch1": "BW01972A", "SecA_Batch2": "BW01972A"}

    claims_dir = output_root / "claims"
    assert sorted(p.name for p in claims_dir.iterdir()) == sorted(["BW01972A", "processing_log.xlsx"])
    shared_file = openpyxl.load_workbook(claims_dir / "BW01972A" / "Claims_SharedSection__BW01972A.xlsx")
    assert set(shared_file.sheetnames) == {"SecA_Batch1", "SecA_Batch2"}


def test_rerun_overwrites_deterministically_and_never_touches_source(tmp_path, monkeypatch):
    entry = FileListEntry(source_file=TEST_FILES_DIR / "Hardy_Risk_test.xlsx", ingestion_type="risk")
    source_bytes_before = entry.source_file.read_bytes()

    output_root_1, _, _ = _run(monkeypatch, tmp_path, [entry])
    output_file = output_root_1 / "premium" / "BW01972" / "Hardy_Risk_test__BW01972.xlsx"

    output_root_2, _, _ = _run(monkeypatch, tmp_path, [entry])

    # same path reused (overwrite, not versioned), and source file untouched
    assert output_root_1 == output_root_2
    assert entry.source_file.read_bytes() == source_bytes_before

    workbook = openpyxl.load_workbook(output_file)
    assert set(workbook.sheetnames) == {"Risk_Main", "Risk_AltAnchor"}


def test_claims_worksheet_with_date_in_name_gets_its_own_file(tmp_path, monkeypatch):
    # Synthetic fixture: one sheet named with a date, one without - exercises the
    # split-by-date-in-name grouping rule not present in the real sample workbooks.
    source_path = tmp_path / "input" / "Claims_Synthetic.xlsx"
    source_path.parent.mkdir(parents=True)
    workbook = openpyxl.Workbook()
    workbook.remove(workbook.active)

    dated_sheet = workbook.create_sheet("Claims_2026-07")
    dated_sheet.append(["Source", "Test"])
    dated_sheet.append(["Claim Number", "Claim Amount"])
    dated_sheet.append(["CLM-1", 100])
    dated_sheet.append(["CLM-2", 200])

    no_date_sheet = workbook.create_sheet("Claims_NoDate")
    no_date_sheet.append(["Source", "Test"])
    no_date_sheet.append(["Claim Number", "Claim Amount"])
    no_date_sheet.append(["CLM-3", 300])

    workbook.save(source_path)

    entry = FileListEntry(source_file=source_path, ingestion_type="claims")
    output_root, _, _ = _run(monkeypatch, tmp_path, [entry])

    claims_dir = output_root / "claims"
    own_file = claims_dir / "Claims_Synthetic__Claims_2026-07.xlsx"
    combined_file = claims_dir / "Claims_Synthetic.xlsx"

    assert own_file.exists()
    assert combined_file.exists()

    own_workbook = openpyxl.load_workbook(own_file)
    assert own_workbook.sheetnames == ["Claims_2026-07"]

    combined_workbook = openpyxl.load_workbook(combined_file)
    assert combined_workbook.sheetnames == ["Claims_NoDate"]


def test_processing_log_xlsx_written_to_each_output_folder(tmp_path, monkeypatch):
    risk_entry = FileListEntry(source_file=TEST_FILES_DIR / "Hardy_Risk_test.xlsx", ingestion_type="risk")
    claims_entry = FileListEntry(source_file=TEST_FILES_DIR / "HDI_Claims_test.xlsx", ingestion_type="claims")

    output_root, _, _ = _run(monkeypatch, tmp_path, [risk_entry, claims_entry])

    risk_log = openpyxl.load_workbook(output_root / "risk" / "processing_log.xlsx")
    claims_log = openpyxl.load_workbook(output_root / "claims" / "processing_log.xlsx")

    assert risk_log.sheetnames == ["Granular", "Overview", "ColumnMapping"]
    assert claims_log.sheetnames == ["Granular", "Overview", "ColumnMapping"]

    # risk's log must not contain claims worksheets and vice versa
    risk_sheet_names = [row[1].value for row in risk_log["Granular"].iter_rows(min_row=2)]
    assert set(risk_sheet_names) == {"Risk_Main", "Risk_AltAnchor", "Risk_NoData", "Notes"}
    claims_sheet_names = [row[1].value for row in claims_log["Granular"].iter_rows(min_row=2)]
    assert set(claims_sheet_names) == {"Claims_Main", "Claims_Secondary", "Claims_NoData", "ReadMe"}


def test_processing_log_never_written_for_a_type_with_no_files_processed(tmp_path, monkeypatch):
    entry = FileListEntry(source_file=TEST_FILES_DIR / "Hardy_Risk_test.xlsx", ingestion_type="risk")

    output_root, _, _ = _run(monkeypatch, tmp_path, [entry])

    assert not (output_root / "claims").exists()


def test_processing_log_csv_persists_and_accumulates_across_runs(tmp_path, monkeypatch):
    entry = FileListEntry(source_file=TEST_FILES_DIR / "Hardy_Risk_test.xlsx", ingestion_type="risk")

    _, csv_path_1, _ = _run(monkeypatch, tmp_path, [entry])
    _, csv_path_2, _ = _run(monkeypatch, tmp_path, [entry])

    assert csv_path_1 == csv_path_2
    rows = _read_csv_rows(csv_path_1)
    run_ids = {r["run_id"] for r in rows if r["step"] == "run_start"}
    assert len(run_ids) == 2  # both runs' evidence accumulated in the same file
